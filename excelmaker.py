from openpyxl import load_workbook
from datetime import datetime


def make_excel(path) -> tuple[float, str, float, str, float]:
    filename: str = path.split("/")[-1]  # exract filename from path
    wb = load_workbook("TempTemplate.xlsx")
    ws = wb.active
    # ws.title = filename
    datum: list[int] = [int(x) for x in filename.split("-")]
    maxTemp: float = 0
    maxTempCas: str = ""
    minTemp: float = 100
    minTempCas: str = ""
    kdy: list[int] = []
    teploty: list[float] = []
    with open(path, "r") as f:
        for i, line in enumerate(f):
            if i != 0:
                for j, item in enumerate(line.split(";")):
                    if j == 0:
                        kdy = [int(x) for x in item.split(":")]
                        ws.cell(row=i + 1, column=j + 1, value=datetime(*datum, *kdy))  # type: ignore # nevím, prostě se tomu něco nelíbí
                    elif j == 1:
                        teplota: float = float(item.replace(",", "."))
                        teploty.append(teplota)
                        ws.cell(row=i + 1, column=j + 1, value=teplota)
                        if teplota > maxTemp:
                            maxTemp = teplota
                            maxTempCas = ":".join([str(x) for x in kdy])
                        if teplota < minTemp:
                            minTemp = teplota
                            minTempCas = ":".join([str(x) for x in kdy])

    avg: float = round(sum(teploty) / len(teploty), 2)
    wb.save(f"{path}.xlsx")
    return maxTemp, maxTempCas, minTemp, minTempCas, avg


if __name__ == "__main__":
    make_excel("WIP/2022-12-04")
